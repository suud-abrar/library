from flask import Flask, render_template,request, redirect
import sqlite3
import datetime
import openpyxl
from twilio.rest import Client

auth_token="54438ef2820e9aa16a01ddb030052468"
account_sid="AC1311a310fd9f01456f3063e7f79c800a"


wb=openpyxl.load_workbook("C:\\Users\\Win 10 Pro\\PycharmProjects\\libraryProject\\project\\tabledata.xlsx")
sh1=wb['newuser']
sh2=wb['borrow']
sh3=wb['online']


app = Flask(__name__)

def get_db_connection():
    conn = sqlite3.connect('database.db')
    conn.row_factory = sqlite3.Row
    return conn
'''
@app.route('/')
def accss():
  connection = get_db_connection()
  cursor = connection.cursor()

  # Replace 'your_table_name' and 'column1', 'column2' with your actual table and columns
  cursor.execute("SELECT username, password FROM signup")
  rows = cursor.fetchall()

  connection.close()
  print(rows[0]['username'])
  for row in rows:
      print(f"Username: {row['username']}, password: {row['password']}")  # Assuming username and email columns
      return
'''

def create_table():
    conn = get_db_connection()
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS signup (username text, password text)''')
    conn.commit()
    conn.close()

# Route to handle form submission
@app.route('/addrec', methods=['POST'])
def addrec():
    name = request.form['username']
    pass1 = request.form['pass']
    makeconn = get_db_connection()
    cursor = makeconn.cursor()
    cursor.execute("SELECT username, password FROM signup")
    rows = cursor.fetchall()
    new=True
    for row in rows:
        if row['username']==name:
            new=False
            break
    if new:
        cursor.execute("INSERT INTO signup (username, password) VALUES (?, ?)", (name, pass1))
        makeconn.commit()
        makeconn.close()
        return "Data submitted successfully!"
    else:
        return ("This username is already taken by other user please try with other username"
                " Or if you already create account please,try to signin")
@app.route('/borrow', methods=['POST'])
def borrow():
    id1 = request.form['studentID']
    grade = request.form['bookGrade']
    type1 = request.form['bookType']
    subject = request.form['subject']
    time = request.form['timeSpent']
    pass2=request.form['password']
    print(id,grade,type1)
    return1 = ""
    rnum = 0
    rnum2 = 0
    row = sh1.max_row
    row2 = sh2.max_row
    row3 = sh3.max_row
    for i in range(row):
        if id1 == sh1.cell(i + 1, 3).value:
            rnum = i + 1
            passwordis = sh1.cell(rnum, 5).value
            break
    for i in range(row3):
        if id1 == sh3.cell(i + 1, 4).value:
            sh3.delete_rows(i + 1)
            break
    for i in range(row2):
        if "yes" != str(sh2.cell(i + 1, 10).value) and str(sh2.cell(i + 1, 4).value) == id1:
            rnum2 = i + 1
            break
    print(passwordis)
    print(pass2)
    if (rnum != 0) and (rnum2 == 0) and int(time) <= 5 and int(time) >= 1 and passwordis==pass2:
        num = str(sh3.max_row) + "."
        fnameof = sh1.cell(rnum, 1).value
        lnameof = sh1.cell(rnum, 2).value
        phoneN = sh1.cell(rnum, 4).value
        currentDT = datetime.datetime.now()
        row2 = sh2.max_row
        loandata = [num, fnameof, lnameof, id1, grade, subject, type1,
                    str(currentDT.date())]
        row3 = sh3.max_row
        for i in range(len(loandata)):

            sh3.cell(row3 + 1, i + 1, value=loandata[i])
        wb.save("C:\\Users\\Win 10 Pro\\PycharmProjects\\libraryProject\\project\\tabledata.xlsx")
        currentDT = datetime.datetime.now()
        print("The proccesse is successfully completed")
        clien = Client(account_sid, auth_token)
        massage = clien.messages.create(
            body="Hi " + fnameof + " " + lnameof + ", Your request is submitted successfully you can take the book at any time " + str(
                currentDT.date()),
            from_="+13343731074",
            to="+251947853943"
        )
        return1 = "The proccesse is successfully completed"
    else:
        if rnum == 0:
            print("This user not belong the system")
            return1 = "This user not belong the system"
        elif rnum2 != 0:
            print("you have already taken a book before, please return it")
            return1 = "you have already taken a book before, please return it"
        elif passwordis!=pass2:
            return "Incorrect password"
        else:
            return1 = "beyond the maximum time"
            print("beyond the maximum time")
    return return1
@app.route('/TOHOME', methods=['POST'])
def TOHOME():
    NAME = request.form['username1']
    PASS = request.form['pass1']
    makeconn = get_db_connection()
    cursor = makeconn.cursor()
    cursor.execute("SELECT username, password FROM signup")
    rows = cursor.fetchall()
    for row in rows:
        if row['username']==NAME and row['password']==PASS:
            return render_template('home.html')
            break
        else:
            if row['username']!=NAME:
                return 'This user name is not found, please cheach your spelling or signup.'
            else:
                return 'Incorrect password'
@app.route('/')
def index():
    return render_template('signup.html')
@app.route('/index11')
def index11():
    return render_template('signup.html')
@app.route('/index1')
def index1():
    return render_template('signin.html')
@app.route('/index2')
def index2():
    return render_template('home.html')
@app.route('/borrowpage')
def borrowpage():
    return render_template('borrow.html')

if __name__ == '__main__':
    app.run(debug=True)
