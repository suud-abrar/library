import customtkinter
from project import dashboard, addcustemer, Returne,borrow,History,online
import openpyxl
from PIL import Image
import customtkinter as ctk
import tkinter as tk
import openpyxl
wb=openpyxl.load_workbook("tabledata.xlsx")
sh1=wb['online']
numberofrow=sh1.max_row-1

wb=openpyxl.load_workbook("tabledata.xlsx")
sh1=wb['adminster']


class Dashboard:
    def __init__(self,window):
        self.window = window
        self.window.title('Library Management')
        self.window.attributes('-fullscreen',True)
        #self.window.geometry('1366x768')
        self.window.state('zoomed')
        self.window.config(background='#673F80')
        #window icon
        icon=customtkinter.CTkImage(light_image=Image.open('images\\historyicon.png'),
                               dark_image=Image.open('images\\historyicon.png'),
                               size=(27, 27))
        #icon=PhotoImage(file='images\\\\\\\\pageicon.svg')
        #self.major_frameiconphoto(True,icon)

        self.major_frame = customtkinter.CTkFrame(self.window, fg_color="#673F80", bg_color="#673F80")
        self.major_frame.pack()
        self.major_frame.pack_propagate(False)
        self.major_frame.configure(width=1366, height=768)


        def delete_page1():
            for frame in self.major_frame.winfo_children():
                frame.destroy()

        def indicator1(path):
            delete_page1()
            path(self.major_frame)
        def delete_page():
            for frame in self.main_frame.winfo_children():
                frame.destroy()

        def indicator(path):
            delete_page()
            path()
        #====================sighnin================
        def sighninin(main_frame):
            def checker():
                rnum = 0
                numrow=0
                row = sh1.max_row
                for i in range(row-1):
                    if input1.get() == str(sh1.cell(i + 1, 1).value):
                        print("hi")
                        rnum = i + 1
                        numrow=i
                        break
                if rnum != 0:
                    if input2.get() == str(sh1.cell(numrow+1, 2).value):
                        indicator1(mainpage1)
                    else:
                        input2.configure(text_color="red")
                        input2.delete(0, 40)
                        input2.insert(0, "Incorrect password")
                else:
                    input1.configure(text_color="red")
                    input1.delete(0, 40)
                    input1.insert(0, "Incorrect ID")

            imagebg = customtkinter.CTkImage(light_image=Image.open('images\\bg.jpg'),
                                             dark_image=Image.open('images\\bg.jpg'),
                                             size=(1366, 768))
            lablebg = customtkinter.CTkLabel(main_frame, image=imagebg)
            lablebg.place(x=0, y=0)
            block1 = customtkinter.CTkFrame(main_frame, border_width=0, border_color="red", fg_color="#412948",
                                            bg_color="#673F80", corner_radius=(10), width=585, height=680)
            block1.place(x=102, y=44)
            head = customtkinter.CTkLabel(main_frame, text="Bookworm ", font=("Poppins", 40, 'bold'),
                                          fg_color="#412948")
            head.place(x=285, y=165)
            imagename = customtkinter.CTkImage(light_image=Image.open('images\\logofrontimage.png'),
                                               dark_image=Image.open('images\\logofrontimage.png'),
                                               size=(188, 209))
            lableimage = customtkinter.CTkLabel(main_frame, text="", image=imagename, fg_color="#412948")
            lableimage.place(x=274, y=250)
            moti = customtkinter.CTkLabel(main_frame, text="For All Patron", font=("Poppins", 30, 'normal'),
                                          fg_color="#412948")
            moti.place(x=290, y=453)
            moti1 = customtkinter.CTkLabel(main_frame, text="Dive into a book today and ",
                                           font=("StalinistOne", 18, 'normal'), bg_color='#412948', fg_color="#412948")
            moti1.place(x=270, y=544)
            moti2 = customtkinter.CTkLabel(main_frame, text="embark on a journey of endless possibilities",
                                           font=("StalinistOne", 18, 'normal'), bg_color='#412948', fg_color="#412948")
            moti2.place(x=220, y=570)
            block2 = customtkinter.CTkFrame(main_frame, border_width=0, border_color="red", fg_color="#900C3F",
                                            bg_color="#673F80",
                                            corner_radius=10, width=577, height=680)
            block2.place(x=687, y=44)
            head1 = customtkinter.CTkLabel(main_frame, text="Welcome Back", font=("Poppins", 33, 'bold'),
                                           fg_color="#900C3F")
            head1.place(x=860, y=137)
            head2 = customtkinter.CTkLabel(main_frame, text="Sign In", font=("Poppins", 33, 'bold'), fg_color="#900C3F")
            head2.place(x=909, y=212)

            label = customtkinter.CTkLabel(main_frame, text="Username")

            def click1(event):
                input1.delete(0, 10)
            def click2(event):
                input2.delete(0, 10)
            varinput1 = customtkinter.StringVar()
            input1 = customtkinter.CTkEntry(main_frame, width=350, height=40, fg_color="#900C3F", bg_color="#900C3F",
                                            border_width=1, border_color="white", placeholder_text="Here Your ID",textvariable=varinput1)
            input1.place(x=687 + 110, y=44 + 245)
            input1.insert(0, "ID")
            input1.bind("<Button-1>", click1)
            varinput2 = customtkinter.StringVar()
            input2 = customtkinter.CTkEntry(main_frame, width=350, height=40, fg_color="#900C3F", bg_color="#900C3F",
                                            border_width=1, border_color="white", placeholder_text="Here Your Password",textvariable=varinput2)
            input2.place(x=687 + 110, y=44 + 338)
            input2.insert(0, "Password")
            input2.bind("<Button-1>", click2)
            head3 = customtkinter.CTkLabel(main_frame, text="forgot password? ", font=("Poppins", 18, 'normal'),
                                           fg_color="#900C3F",bg_color="#900C3F")
            head3.place(x=687 + 170, y=44 + 531)
            head4 = customtkinter.CTkLabel(main_frame, text=" click here", font=("Poppins", 18, 'normal'),
                                           fg_color="#900C3F",bg_color="#900C3F", text_color="red")
            head4.place(x=687 + 315, y=44 + 531)
            logbutton = customtkinter.CTkButton(main_frame, text='Sign In', font=("Poppins", 33, "normal"),
                                                border_width=0, hover_color="#0B7515"
                                                , cursor="hand2", fg_color="#1D9321", bg_color="#900C3F", width=347,
                                                height=65,command=lambda:checker())
            logbutton.place(x=687 + 114, y=44 + 438)
        #===================main page================

        def mainpage1(major_frame):
            self.main_frame = customtkinter.CTkFrame(major_frame, fg_color="#673F80", bg_color="#673F80")
            self.main_frame.pack(side="right")
            self.main_frame.pack_propagate(False)
            self.main_frame.configure(width=1366 - 238, height=768)
            def change1():
                self.dashboardbutton.configure(fg_color='#A569BD')
                self.borrowbutton.configure(fg_color='#412948')
                self.reternbutton.configure(fg_color='#412948')
                self.newuserbutton.configure(fg_color='#412948')
                self.historybutton.configure(fg_color='#412948')
                self.logoutbutton.configure(fg_color='#412948')
            def dashboardfun():
                dashboard.dashboardin(self.main_frame)
            dashboardfun()
            def borrowfun():
                borrow.borrowin(self.main_frame)
            def reternfun():
                Returne.reternin(self.main_frame)
            def addfun():
                addcustemer.addin(self.main_frame)
            def historyfun():
                History.historyin(self.main_frame)
            def onlinefun():
                online.onlinein(self.main_frame)
            # ===================SIDEBAR====================
            self.bar = customtkinter.CTkFrame(self.major_frame, fg_color='#412948', width=238, height=768)
            self.bar.place(x=0, y=0)

            self.imagename = customtkinter.CTkImage(light_image=Image.open('images\\logoimge.png'),
                                                    dark_image=Image.open('images\\logoimge.png'),
                                                    size=(129, 129))
            self.lableimage = customtkinter.CTkLabel(self.major_frame, text="", image=self.imagename, fg_color="#412948")
            self.lableimage.place(x=42, y=30)

            # ===================barbutton=======================
            self.framebt1=customtkinter.CTkFrame(self.major_frame)
            self.framebt1.place()


            self.dashboardbutton = customtkinter.CTkButton(self.major_frame, text='Dashboard', bg_color='#673F80',
                                                           font=("Poppins", 22, "normal"),
                                                           fg_color='#A569BD', cursor='hand2', width=224, height=53,
                                                           border_width=0,hover_color='#A055B4', command=lambda: indicator(dashboardfun))
            self.dashboardbutton.place(x=7, y=197)

            self.darboardicon = customtkinter.CTkImage(light_image=Image.open('images\\dash.png'),
                                                       dark_image=Image.open('images\\dash.png'),
                                                       size=(27, 27))
            self.labledash = customtkinter.CTkLabel(self.major_frame, text="", image=self.darboardicon, fg_color="#A569BD")
            self.labledash.place(x=31, y=207)


            self.borrowbutton = customtkinter.CTkButton(self.major_frame, text='Borrow', bg_color='#412948',
                                                        font=("Poppins", 22, "normal"),
                                                        fg_color='#A569BD', cursor='hand2', width=224, height=53,
                                                        border_width=0,hover_color='#A055B4', command=lambda: indicator(borrowfun))
            self.borrowbutton.place(x=7, y=281)
            self.borrowicon = customtkinter.CTkImage(light_image=Image.open('images\\borrow.png'),
                                                     dark_image=Image.open('images\\borrow.png'),
                                                     size=(33, 25))
            self.lableborrow = customtkinter.CTkLabel(self.major_frame, text="", image=self.borrowicon, fg_color="#A569BD")
            self.lableborrow.place(x=32, y=290)

            self.reternbutton = customtkinter.CTkButton(self.major_frame, text='Returner', bg_color='#412948',
                                                        font=("Poppins", 22, "normal"),
                                                        fg_color='#A569BD', cursor='hand2', width=224, height=53,
                                                        border_width=0,hover_color='#A055B4', command=lambda: indicator(reternfun))
            self.reternbutton.place(x=7, y=368)
            self.reternicon = customtkinter.CTkImage(light_image=Image.open('images\\retern.png'),
                                                     dark_image=Image.open('images\\retern.png'),
                                                     size=(25, 25))
            self.lableretern = customtkinter.CTkLabel(self.major_frame, text="", image=self.reternicon, fg_color="#A569BD")
            self.lableretern.place(x=32, y=377)

            self.newuserbutton = customtkinter.CTkButton(self.major_frame, text='New User', bg_color='#412948',
                                                         font=("Poppins", 22, "normal"),
                                                         fg_color='#A569BD', cursor='hand2', width=224, height=53,
                                                         border_width=0,hover_color='#A055B4', command=lambda: indicator(addfun))
            self.newuserbutton.place(x=7, y=449)
            self.newusericon = customtkinter.CTkImage(light_image=Image.open('images\\newuser.png'),
                                                      dark_image=Image.open('images\\newuser.png'),
                                                      size=(38, 22))
            self.lablenewuser = customtkinter.CTkLabel(self.major_frame, text="", image=self.newusericon, fg_color="#A569BD")
            self.lablenewuser.place(x=25, y=459)

            self.historybutton = customtkinter.CTkButton(self.major_frame, text='History', bg_color='#412948',
                                                         font=("Poppins", 22, "normal"),
                                                         fg_color='#A569BD', cursor='hand2', width=224, height=53,
                                                         border_width=0,hover_color='#A055B4',command=lambda: indicator(historyfun))
            self.historybutton.place(x=7, y=533)
            self.historyicon = customtkinter.CTkImage(light_image=Image.open('images\\historyicon.png'),
                                                      dark_image=Image.open('images\\historyicon.png'),
                                                      size=(27, 27))
            self.lablehistory = customtkinter.CTkLabel(self.major_frame, text="", image=self.historyicon, fg_color="#A569BD")
            self.lablehistory.place(x=28, y=543)
            self.onlinebtn = customtkinter.CTkButton(self.major_frame, text='Online', bg_color='#412948',
                                                         font=("Poppins", 22, "normal"),
                                                         fg_color='#A569BD', cursor='hand2', width=224, height=53,
                                                         border_width=0, hover_color='#A055B4',
                                                         command=lambda: indicator(onlinefun))
            self.onlinebtn.place(x=7, y=533+80)
            self.onlineicon = customtkinter.CTkImage(light_image=Image.open('images\\onlineicon.png'),
                                                      dark_image=Image.open('images\\onlineicon.png'),
                                                      size=(27, 27),)
            self.olinelable = customtkinter.CTkLabel(self.major_frame, text="", image=self.onlineicon,
                                                       fg_color="#A569BD")
            self.olinelable.place(x=28, y=543+80)

            # ===============line================
            self.frameline = customtkinter.CTkFrame(self.major_frame, border_width=1, corner_radius=1, width=183, height=1,
                                                    fg_color="white", bg_color="#673F80")
            self.frameline.place(x=24, y=673)

            self.logoutbutton = customtkinter.CTkButton(self.major_frame, text='Logout', bg_color='#412948',
                                                        font=("Poppins", 22, "normal"),
                                                        fg_color='#A569BD', cursor='hand2', width=125, height=46,
                                                        border_width=0,hover_color='#A055B4',command=lambda: indicator1(sighninin))
            self.logoutbutton.place(x=40, y=704)
            self.logouticon = customtkinter.CTkImage(light_image=Image.open('images\\logouticon.png'),
                                                     dark_image=Image.open('images\\logouticon.png'),
                                                     size=(27, 24))
            self.lablelogout = customtkinter.CTkLabel(self.major_frame, text="", image=self.logouticon, fg_color="#A569BD")
            self.lablelogout.place(x=40, y=712)

            #==========================exit or quit================
            setting = customtkinter.CTkButton(self.major_frame, bg_color='#412948', width=25, height=25, text='Exit',
                                             command=window.quit,border_width=0,hover_color='#A055B4' ,fg_color='#A569BD', cursor='hand2')
            setting.place(x=1307, y=40)
            canvas = tk.Canvas(self.major_frame, bg="white")
            canvas.configure(width=20,height=20)
            canvas.place(x=170,y=626)
            notiflable=customtkinter.CTkLabel(self.major_frame,text=numberofrow,font=("Poppins", 20, "normal"),
                                              text_color='black' ,height=3,bg_color='white',fg_color='white')
            notiflable.place(x=177,y=627)

        #indicator1(sighninin)
        indicator1(mainpage1)

def win():
    window = customtkinter.CTk()
    Dashboard(window)
    window.mainloop()
if __name__=='__main__':
    win()