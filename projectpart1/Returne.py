import customtkinter
import tkinter as tk
from tkinter import *
from PIL import Image, ImageTk
from tkinter import ttk
import openpyxl

wb=openpyxl.load_workbook("tabledata.xlsx")
sh1=wb['borrow']
sheet1=wb['hoster']
x=33
y=85
def reternin(main_frame):
    # ===================header1====================
    header = customtkinter.CTkLabel(main_frame, text="Return", font=("Poppins", 29, 'bold'),
                                    bg_color='#673F80', fg_color='#673F80', width=199, height=55)
    header.place(x=474, y=21)
    frameline = customtkinter.CTkFrame(main_frame, border_width=1, corner_radius=1, width=1060, height=1,
                                       fg_color="white", bg_color="#673F80")
    frameline.place(x=33, y=78)
    presecondframe = customtkinter.CTkFrame(main_frame, border_width=0, corner_radius=1, width=1060, height=375,
                                            fg_color="#673F80", bg_color="#673F80")
    presecondframe.place(x=33, y=81)

    def tablefun():
        def load_data():
            path = "tabledata.xlsx"
            workbook = openpyxl.load_workbook(path)
            sh1 = workbook['borrow']
            list_value = list(sh1.values)
            for col_name in list_value[0]:
                Tview.heading(col_name, text=col_name)
            for value_tuple in list_value[1:]:

                if value_tuple[9] != "yes":
                    Tview.insert('', tk.END, values=value_tuple)

        secondframe = customtkinter.CTkFrame(presecondframe,fg_color="white")
        secondframe.place(x=2, y=2)
        scrool = customtkinter.CTkScrollbar(secondframe)
        scrool.pack(side="right", fill="y")
        style = ttk.Style()

        style.theme_use("clam")
        style.configure("Treeview",
                        font=("Calibre", 15),
                        background="#412948",
                        foreground="white",
                        rowheight=56,
                        fieldbackground="#412948")
        style.map('Treeview',
                  background=[('selected', '#A569BD')])

        cols = ["No.", "First Name", "Last Name", "ID", "Grade", "Subject", "Type", "Time", "return on", "returned"]
        Tview = ttk.Treeview(secondframe, show="headings", yscrollcommand=scrool.set, columns=cols, height=6)
        Tview.column("No.", width=38)
        Tview.column("First Name", width=136)
        Tview.column("Last Name", width=136)
        Tview.column("ID", width=66)
        Tview.column("Grade", width=96)
        Tview.column("Subject", width=120)
        Tview.column("Type", width=87)
        Tview.column("Time", width=155)
        Tview.column("return on", width=155)
        Tview.column("returned", width=52)
        Tview.pack()
        scrool.configure(command=Tview.yview)
        load_data()
        def line(x,y):
            frameline = customtkinter.CTkFrame(secondframe, border_width=1, corner_radius=1, width=1030, height=1,
                                               fg_color="white", bg_color="#673F80")
            frameline.place(x=x, y=y)
        line(3,75)
        line(3, 75 + 56)
        line(3,75+56+56)
        line(3,75+56+56+56)
        line(3,75+56+56+56+56)
        line(3,75+56+56+56+56+56)

    tablefun()

    def delete_page():
        for frame in presecondframe.winfo_children():
            frame.destroy()

    def indicator(path):
        delete_page()
        path()

    def FIND():
        def upperfun():
            rnum = 0
            row = sh1.max_row
            for i in range(row):
                if inputret.get() == str(sh1.cell(i + 1, 4).value) and sh1.cell(i + 1, 10).value != 'yes':
                    apperframe = customtkinter.CTkFrame(presecondframe, border_width=0, corner_radius=15, width=1060,
                                                        height=262,
                                                        fg_color="#412948", bg_color="#673F80")
                    apperframe.place(x=0, y=10)
                    rnum = i + 1
                    FN = sh1.cell(rnum, 2).value
                    LN = sh1.cell(rnum, 3).value
                    IDS = sh1.cell(rnum, 4).value
                    GS = sh1.cell(rnum, 5).value
                    SS = sh1.cell(rnum, 6).value
                    TP = sh1.cell(rnum, 7).value
                    TIME = sh1.cell(rnum, 8).value
                    RETTIME = sh1.cell(rnum, 9).value

                    def upperlable(texts, x1, y1):
                        lable = customtkinter.CTkLabel(apperframe, text=texts,
                                                       font=("Poppins", 22, "normal"), width=136,
                                                       height=15, bg_color="#412948", fg_color="#412948",
                                                       )
                        lable.place(x=x1, y=y1)

                    upperlable("first name", 1, 19)
                    upperlable("last name", 136, 19)
                    upperlable("id", 136 + 136, 19)
                    upperlable("grade", 136 + 136 + 136, 19)
                    upperlable("subject", 136 + 136 + 136 + 136, 19)
                    upperlable("type", 136 + 136 + 136 + 136 + 136, 19)
                    upperlable("time", 136 + 136 + 136 + 136 + 136 + 100, 19)
                    upperlable("ret_time", 136 + 136 + 136 + 136 + 136 + 136 + 100, 19)
                    frameline = customtkinter.CTkFrame(apperframe, border_width=1, corner_radius=1, width=1012,
                                                       height=1,
                                                       fg_color="white", bg_color="#673F80")
                    frameline.place(x=15, y=55)
                    upperlable(FN, 1, 60)
                    upperlable(LN, 136, 60)
                    upperlable(IDS, 136 + 136, 60)
                    upperlable(GS, 136 + 136 + 136, 60)
                    upperlable(SS, 136 + 136 + 136 + 136, 60)
                    upperlable(TP, 136 + 136 + 136 + 136 + 136, 60)
                    upperlable(TIME, 136 + 136 + 136 + 136 + 136 + 100, 60)
                    upperlable(RETTIME, 136 + 136 + 136 + 136 + 136 + 136 + 100, 60)
                    idofs = sh1.cell(i + 1, 4).value

                    def RETERN():
                        sh1.cell(rnum, 10, 'yes')
                        rownum = sheet1.max_row
                        for i in range(rownum):
                            if idofs==sheet1.cell(i+1,1):
                                sheet1.delete_rows(i+1)
                                wb.save("tabledata.xlsx")
                        wb.save("tabledata.xlsx")
                        indicator(tablefun)

                    reternbt = customtkinter.CTkButton(apperframe, text='RETERN', bg_color='#412948',
                                                       fg_color='#1D9321',
                                                       font=("Poppins", 20, "normal"),
                                                       cursor='hand2', hover_color="#0B7515", command=RETERN)
                    reternbt.place(x=442, y=131)
                    break
        indicator(upperfun)
        inputret.delete(0, 10)
        inputret.insert(0, "ID")

    def click1(event):
        inputret.delete(0, 10)
    lowerframe = customtkinter.CTkFrame(main_frame, border_width=0, corner_radius=15, width=1060, height=262,
                                        fg_color="#412948", bg_color="#673F80")
    lowerframe.place(x=33, y=475)
    lable1 = customtkinter.CTkLabel(lowerframe, text="Find by ID:", font=("Poppins", 30, "normal"), width=151,
                                    height=36, bg_color="#412948", fg_color="#412948")
    lable1.place(x=175, y=40)

    varinput = customtkinter.StringVar()
    inputret = customtkinter.CTkEntry(lowerframe, width=252, height=49, fg_color="#412948",
                                      bg_color="#412948",
                                      border_width=1, border_color="white", placeholder_text="ID",
                                      textvariable=varinput)
    inputret.insert(0, "ID")
    inputret.bind("<Button-1>", click1)
    inputret.place(x=124, y=104)

    find = customtkinter.CTkButton(lowerframe, text='Find', bg_color='#412948', fg_color='#1D9321',
                                   font=("Poppins", 20, "normal"),
                                   cursor='hand2', hover_color="#0B7515", command=FIND)
    find.place(x=186, y=181)
    frameline = customtkinter.CTkFrame(lowerframe, border_width=1, corner_radius=1, width=1, height=262 - 2 * 29,
                                       fg_color="white", bg_color="#412948")
    frameline.place(x=523, y=29)
    lable2 = customtkinter.CTkLabel(lowerframe, text="Reading opens doors to new worlds, ",
                                    font=("Poppins", 30, "normal"), width=517,
                                    height=36, bg_color="#412948", fg_color="#412948", text_color="#F39C12")
    lable2.place(x=535, y=40)
    lable3 = customtkinter.CTkLabel(lowerframe, text="expands our minds, and",
                                    font=("Poppins", 30, "normal"), width=339,
                                    height=36, bg_color="#412948", fg_color="#412948", text_color="#F39C12")
    lable3.place(x=693, y=93)
    lable4 = customtkinter.CTkLabel(lowerframe, text="fuels our imagination",
                                    font=("Poppins", 30, "normal"), width=298,
                                    height=36, bg_color="#412948", fg_color="#412948", text_color="#F39C12")
    lable4.place(x=646, y=148)

