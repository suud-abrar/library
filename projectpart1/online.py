import customtkinter
import tkinter as tk
from tkinter import *
from PIL import Image, ImageTk
from tkinter import ttk
import openpyxl
import datetime
from twilio.rest import Client


auth_token="54438ef2820e9aa16a01ddb030052468"
account_sid="AC1311a310fd9f01456f3063e7f79c800a"
wb=openpyxl.load_workbook("tabledata.xlsx")
sh1=wb['online']
sheet1=wb['hoster']
numberofrow=sheet1.max_row


def onlinein(main_frame):
    # ===================header1====================
    header = customtkinter.CTkLabel(main_frame, text="Online", font=("Poppins", 29, 'bold'),
                                    bg_color='#673F80', fg_color='#673F80', width=199, height=55)

    header.place(x=488, y=21)
    frameline = customtkinter.CTkFrame(main_frame, border_width=1, corner_radius=1, width=1060, height=1,
                                       fg_color="white", bg_color="#673F80")
    frameline.place(x=33, y=78)

    def tablefun():
        def load_data():
            path = "tabledata.xlsx"
            workbook = openpyxl.load_workbook(path)
            sh1 = workbook['online']
            list_value = list(sh1.values)
            for col_name in list_value[0]:
                Tview.heading(col_name, text=col_name)
            for value_tuple in list_value[1:]:
                    Tview.insert('', tk.END, values=value_tuple)

        def click1(event):
            inputid.delete(0, 10)

        def delete_page1():
            for frame in main_frame.winfo_children():
                frame.destroy()

        def indicator1(path):
            delete_page1()
            path()
            # ===================header1====================
            header = customtkinter.CTkLabel(main_frame, text="Online", font=("Poppins", 29, 'bold'),
                                            bg_color='#673F80', fg_color='#673F80', width=199, height=55)

            header.place(x=488, y=21)
            frameline = customtkinter.CTkFrame(main_frame, border_width=1, corner_radius=1, width=1060, height=1,
                                               fg_color="white", bg_color="#673F80")
            frameline.place(x=33, y=78)

        def doing(event):
            inputvalue=inputid.get()
            def upperfun():
                rnum = 0
                row = sh1.max_row
                for i in range(row):
                    ii=i
                    if inputvalue == str(sh1.cell(i + 1, 4).value):
                        apperframe = customtkinter.CTkFrame(main_frame, border_width=0, corner_radius=15,
                                                            width=1060,
                                                            height=120,
                                                            fg_color="#412948", bg_color="#673F80")
                        apperframe.place(x=25, y=130)
                        rnum = i + 1
                        FN = sh1.cell(2, 2).value
                        LN = sh1.cell(rnum, 3).value
                        IDS = sh1.cell(rnum, 4).value
                        GS = sh1.cell(rnum, 5).value
                        SS = sh1.cell(rnum, 6).value
                        TP = sh1.cell(rnum, 7).value
                        TIME = sh1.cell(rnum, 8).value

                        def upperlable(texts, x1, y1):
                            lable = customtkinter.CTkLabel(apperframe, text=texts,
                                                           font=("Poppins", 22, "normal"), width=136,
                                                           height=15, bg_color="#412948", fg_color="#412948",
                                                           )
                            lable.place(x=x1, y=y1)

                        upperlable("first name", 1, 19)
                        upperlable("last name", 136, 19)
                        upperlable("id", 136 + 136, 19)
                        upperlable("grade", 136 + 136 + 136, 19)
                        upperlable("subject", 136 + 136 + 136 + 136, 19)
                        upperlable("type", 136 + 136 + 136 + 136 + 136, 19)
                        upperlable("time", 136 + 136 + 136 + 136 + 136 + 100, 19)
                        frameline = customtkinter.CTkFrame(apperframe, border_width=1, corner_radius=1, width=1012-136,
                                                           height=1,
                                                           fg_color="white", bg_color="#673F80")
                        frameline.place(x=15, y=55)
                        upperlable(FN, 1, 60)
                        upperlable(LN, 136, 60)
                        upperlable(IDS, 136 + 136, 60)
                        upperlable(GS, 136 + 136 + 136, 60)
                        upperlable(SS, 136 + 136 + 136 + 136, 60)
                        upperlable(TP, 136 + 136 + 136 + 136 + 136, 60)
                        upperlable(TIME, 136 + 136 + 136 + 136 + 136 + 100, 60)
                        def confirme():
                            sh1.delete_rows(ii+1)
                            currentDT = datetime.datetime.now()
                            loaddata = [IDS, TIME, str(currentDT.day)]
                            for i in range(len(loaddata)):
                                sheet1.cell(numberofrow + 1, i + 1, value=loaddata[i])
                            wb.save("tabledata.xlsx")
                            wb.save("tabledata.xlsx")
                            clien = Client(account_sid, auth_token)
                            massage = clien.messages.create(
                                body="Hi " + FN + " " + LN + ", you have succesfully borrowed grade " + GS + " " + SS + " "+ TP + " on " + str(
                                    currentDT.date()) + " you have to return the book on " + str(
                                    currentDT.year) + "-" + str(
                                    currentDT.month) + "-" + str(currentDT.day + TIME),
                                from_="+13343731074",
                                to="+251947853943"
                            )
                            tablefun()


                        confirmebtn = customtkinter.CTkButton(apperframe, text='Confirm', bg_color='#412948',
                                                           fg_color='#EC9410',
                                                           font=("Poppins", 20, "normal"),
                                                           cursor='hand2', hover_color="#0B7515", command=confirme)

                        confirmebtn.place(x=900, y=40)
                        break
                    else:
                        tablefun()

            indicator1(upperfun)
            inputid.delete(0, 10)
            inputid.insert(0, "Find by ID")
        varinput = customtkinter.StringVar()
        inputid = customtkinter.CTkEntry(main_frame, width=252, height=40, fg_color="#412948",
                                          bg_color="#673F80",
                                          border_width=1, corner_radius=30,border_color="white", placeholder_text="ID",
                                          textvariable=varinput)
        inputid.insert(0, "Find by ID")
        inputid.bind("<Button-1>", click1)
        inputid.place(x=830, y=85)
        inputid.bind("<Return>",doing)
        secondframe = customtkinter.CTkFrame(main_frame,fg_color="#412948",corner_radius=10)
        secondframe.place(x=33, y=130)
        scrool = customtkinter.CTkScrollbar(main_frame)
        scrool.pack(side="right", fill="y")
        style = ttk.Style()

        style.theme_use("default")
        style.configure("Treeview",
                        font=("Calibre", 15),
                        background="#412948",
                        foreground="white",
                        rowheight=56,

                        fieldbackground="#412948")
        style.map('Treeview',
                  background=[('selected', '#A569BD')])

        cols = ["No.", "First Name", "Last Name", "ID", "Grade", "Subject", "Type", "Time"]
        Tview = ttk.Treeview(secondframe, show="headings", yscrollcommand=scrool.set, columns=cols, height=10)
        Tview.column("No.", width=38)
        Tview.column("First Name", width=136+30)
        Tview.column("Last Name", width=136+30)
        Tview.column("ID", width=66+30)
        Tview.column("Grade", width=96+30)
        Tview.column("Subject", width=120+30)
        Tview.column("Type", width=87+30)
        Tview.column("Time", width=155+30)
        Tview.pack()
        scrool.configure(command=Tview.yview)
        load_data()

        def line(x,y):
            frameline = customtkinter.CTkFrame(secondframe, border_width=1, corner_radius=1, width=1030, height=1,
                                               fg_color="white", bg_color="#673F80")
            frameline.place(x=x, y=y)
        line(3,75)
        line(3, 75+56)
        line(3,75+56+56)
        line(3,75+56+56+56)
        line(3,75+56+56+56+56)
        line(3,75+56+56+56+56+56)
        line(3,75+56+56+56+56+56+56)
        line(3,75+56+56+56+56+56+56+56)
        line(3,75+56+56+56+56+56+56+56+56)
        line(3,75+56+56+56+56+56+56+56+56+56)
        line(3,75+56+56+56+56+56+56+56+56+56+56)
    tablefun()

