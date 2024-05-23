import schedule
import time
import openpyxl
from twilio.rest import Client
import datetime

def do():
    print("i am hosting")
    auth_token = "54438ef2820e9aa16a01ddb030052468"
    account_sid = "AC1311a310fd9f01456f3063e7f79c800a"

    wb = openpyxl.load_workbook("tabledata.xlsx")
    sheet1 = wb['hoster']
    sheet2 = wb['newuser']
    rownum1 = sheet1.max_row
    rownum2 = sheet2.max_row
    for i in range(rownum1):
        currentday = datetime.datetime.now().day
        lastday = sheet1.cell(i + 2, 3).value
        lentime = sheet1.cell(i + 2, 2).value
        ID = sheet1.cell(i + 2, 1).value
        if ID!="None" and lentime <= currentday - lastday:
            for i2 in range(rownum2):
                if sheet2.cell(i2 + 1, 3).value == str(ID):
                    print('hi')
                    phonenum = sheet2.cell(i2 + 1, 4)
                    fnameof = sheet2.cell(i2 + 1, 1)
                    lnameof = sheet2.cell(i2 + 1, 2)
                    clien = Client(account_sid, auth_token)
                    massage = clien.messages.create(
                        body="warning!!! " + str(fnameof) + " " + str(
                            lnameof) + "still you haven't bring the book you took, the time is over",
                        from_="+13343731074",
                        to="+251947853943"
                        # + str(phonenum)[1:]
                    )
                    break


schedule.every(5).seconds.do(do)

while True:
    schedule.run_pending()
    time.sleep(2)

