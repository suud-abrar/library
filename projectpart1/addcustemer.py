import customtkinter
import openpyxl
import numpy as np

from twilio.rest import Client


auth_token="54438ef2820e9aa16a01ddb030052468"
account_sid="AC1311a310fd9f01456f3063e7f79c800a"


wb = openpyxl.load_workbook("tabledata.xlsx")
sh1 = wb['newuser']
sh2 = wb['borrow']
row = sh1.max_row
col = sh1.max_column

def addin(main_frame):
    # ===================header1====================
    header = customtkinter.CTkLabel(main_frame, text="New User", font=("Poppins", 29, 'bold'),
                                    bg_color='#673F80', fg_color='#673F80', width=199, height=55)

    header.place(x=488, y=21)
    frameline = customtkinter.CTkFrame(main_frame, border_width=1, corner_radius=1, width=1060, height=1,
                                       fg_color="white", bg_color="#673F80")
    frameline.place(x=33, y=78)
    # ===================main board====================
    board = customtkinter.CTkFrame(main_frame, fg_color='#412948', bg_color="#673F80", border_width=0,
                                        corner_radius=(15), width=974, height=561)
    board.place(x=79, y=116)
    def lables(texts, widhts, px, py):
        lable1 = customtkinter.CTkLabel(main_frame, text=texts, font=("Poppins", 22, "normal"), width=widhts,
                                        height=38, bg_color="#412948", fg_color="#412948")
        lable1.place(x=px, y=py)

    def discard():
        input6.delete(0,30)
        input6.insert(0, "First Name:")
        input6a.delete(0, 30)
        input6a.insert(0, "Last Name:")
        input7.delete(0, 30)
        input7.insert(0, "ID:")
        input8.delete(0, 30)
        input8.insert(0, "Phone Number:")

    def register():
        chars = [['a', 'b', 'c'], ['d', 'e', 'f'], ['g', 'h', 'i'], ['j', 'k', 'l'], ['m', 'n', 'o'],
                 ['p', 'q', 'r'], ['s', 't', 'u'], ['v', 'w', 'x'], ['y', 'z', '0'], ['1', '2', '3'], ['4', '5', '6'],
                 ['7', '8', '9'], ]
        password_length = 6
        chars = np.array(chars).flatten()
        password = ''.join(np.random.choice(chars, password_length))
        list = [input6.get(), input6a.get(), input7.get(), input8.get(),password]
        for i in range(row):
            if input6.get() == sh1.cell(i + 1, 1).value and input6a.get() == sh1.cell(i + 1, 2).value:
                print("this user aready exist")
                break
            elif i + 1 == row:
                for x in range(row):
                    if input7.get() == sh1.cell(x + 1, 3).value:
                        print("this id aready take by other user")
                        lables("This id aready take by other user", 400, 467, 609)
                        break
                    elif x + 1 == row:
                        for j in range(len(list)):
                            sh1.cell(row + 1, j + 1, value=list[j])
                        wb.save("tabledata.xlsx")
                        clien = Client(account_sid, auth_token)
                        massage = clien.messages.create(
                            body="Hi "+password+" this is your web_password" ,
                            from_="+13343731074",
                            to="+251947853943"
                        )
                        print("succ. reg")
                        lables("The Registration is successfully completed", 400, 467, 609)
                        input6.delete(0, 30)
                        input6.insert(0, "First Name:")
                        input6a.delete(0, 30)
                        input6a.insert(0, "Last Name:")
                        input7.delete(0, 30)
                        input7.insert(0, "ID:")
                        input8.delete(0, 30)
                        input8.insert(0, "+251...")

    lables("First Name:", 127, 324, 169)
    lables("Last Name:", 128, 323, 245)
    lables("ID:", 28, 424, 314)
    lables("Phone Number:", 176, 275, 397)


    def click1(event):
        input6.delete(0, 20)

    def click2(event):
        input6a.delete(0, 20)
    def click3(event):
        input7.delete(0, 20)
    def click4(event):
        input8.delete(0, 20)


    varinput1 = customtkinter.StringVar()
    input6 = customtkinter.CTkEntry(main_frame, width=459, height=44, fg_color="#412948",
                                    bg_color="#412948",
                                    border_width=1, border_color="white"
                                    , textvariable=varinput1)
    input6.insert(0, "First Name")
    input6.bind("<Button-1>", click1)
    input6.place(x=474, y=166)
    varinput2 = customtkinter.StringVar()

    input6a = customtkinter.CTkEntry(main_frame, width=459, height=44, fg_color="#412948",
                                     bg_color="#412948",
                                     border_width=1, border_color="white",
                                     textvariable=varinput2)
    input6a.insert(0, "Last Name")
    input6a.bind("<Button-1>", click2)
    input6a.place(x=474, y=242)

    varinput3 = customtkinter.StringVar()
    input7 = customtkinter.CTkEntry(main_frame, width=459, height=44, fg_color="#412948",
                                       bg_color="#412948",
                                       border_width=1, border_color="white",
                                       textvariable=varinput3)
    input7.insert(0, "ID")
    input7.bind("<Button-1>", click3)
    input7.place(x=474, y=318)

    varinput4 = customtkinter.StringVar()
    input8 = customtkinter.CTkEntry(main_frame, width=459, height=44, fg_color="#412948",
                                    bg_color="#412948",
                                    border_width=1, border_color="white",
                                      textvariable=varinput4)
    input8.insert(0, "+251...")
    input8.bind("<Button-1>", click4)
    input8.place(x=474, y=394)


    reg = customtkinter.CTkButton(main_frame, text="Register",bg_color='#412948', fg_color='#1D9321',
                                        font=("Poppins", 20, "normal"),
                                        cursor='hand2', hover_color="#0B7515", command=register)
    reg.place(x=574, y=540)
    dis = customtkinter.CTkButton(main_frame, text="Discard", bg_color='#412948', fg_color='#D80A12',
                                        font=("Poppins", 20, "normal"),
                                        cursor='hand2', hover_color="#0B7515", command=discard)
    dis.place(x=739, y=540)

