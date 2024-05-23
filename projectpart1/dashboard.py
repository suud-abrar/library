import customtkinter
from PIL import Image, ImageTk
import openpyxl

wb=openpyxl.load_workbook("tabledata.xlsx")
sh9=wb['grade9and10']
sh11=wb['grade11and12']
def dashboardin(main_frame):

    # ===================heading====================
    header = customtkinter.CTkLabel(main_frame, text="Dashboard", font=("Poppins", 29, 'bold'),
                                         bg_color='#673F80', fg_color='#673F80')
    header.place(x=436, y=21)

    frameline = customtkinter.CTkFrame(main_frame, border_width=1, corner_radius=1, width=977, height=2,
                                            fg_color="white", bg_color="#673F80")
    frameline.place(x=44, y=78)
    # ===================SIDEBAR2====================
    bar2 = customtkinter.CTkFrame(main_frame, bg_color='#412948',fg_color='#412948', width=90, height=768)
    bar2.place(x=1276-238, y=0)
    # circle
    settingicon = customtkinter.CTkImage(light_image=Image.open('images\\onlineicon.png'),
                                             dark_image=Image.open('images\\onlineicon.png'),
                                             size=(27, 27), )
    setting = customtkinter.CTkButton(main_frame, bg_color='#412948',width=25,height=25,text='',image=settingicon)
    setting.place(x=1070, y=94)

    # ===================books===================

    EC910 = sh9.cell(4, 2).value
    EM1112 = sh11.cell(2, 2).value
    EP1112 = sh11.cell(5, 2).value
    SM910 = sh9.cell(2, 3).value
    def bookfame(x, y, img, Bnom):
        book = customtkinter.CTkFrame(main_frame, border_width=0, fg_color='#412948', bg_color="#673F80",
                                           width=111, height=139, corner_radius=(8))
        book.place(x=x, y=y)
        imagename = customtkinter.CTkImage(light_image=Image.open('images\\' + img + '.png'),
                                                dark_image=Image.open('images\\' + img + '.png')
                                                , size=(86, 104))
        lableimage = customtkinter.CTkLabel(main_frame, text="", image=imagename, fg_color="#412948")
        lableimage.place(x=x + 14, y=y + 5)
        book_no = customtkinter.CTkLabel(main_frame, text=Bnom, font=("Poppins", 13, 'bold'), width=27, height=6,
                                              bg_color='#412948', fg_color='#412948')
        book_no.place(x=x + 44, y=y + 109)
        book_ = customtkinter.CTkLabel(main_frame, text="books", font=("Poppins", 13, 'bold'),
                                       bg_color='#412948', fg_color='#412948', width=48, height=6)
        book_.place(x=x + 36, y=y + 122)


    y = 102
    for i1 in range(4):
        print(i1)
        x = 52
        for i in range(4):
            if i==0 and i1==0:
                Bnum=EC910

            elif i==1 and i1==0:
                Bnum=EM1112
            elif i==2 and i1==0:
                Bnum=EP1112
            elif i==3 and i1==0:
                Bnum=SM910
            else:
                Bnum = 0
            bookfame(x, y, str(i) + str(i1), str(Bnum))
            x += 148
        y += 162

    def bookinfo(x, y):
        bookinf = customtkinter.CTkFrame(main_frame, border_width=0, fg_color='#412948', bg_color="#673F80",
                                              width=359, height=301, corner_radius=(15))
        bookinf.place(x=x, y=y)

    bookinfo(643, 102)

    bookinfo(643, 426)
