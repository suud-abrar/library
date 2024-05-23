import customtkinter
import datetime
import openpyxl
from twilio.rest import Client


auth_token="54438ef2820e9aa16a01ddb030052468"
account_sid="AC1311a310fd9f01456f3063e7f79c800a"


wb=openpyxl.load_workbook("tabledata.xlsx")
sh1=wb['newuser']
sh2=wb['borrow']
sheet1=wb['hoster']
numberofrow=sheet1.max_row
def borrowin(main_frame):
    # ===================header1====================
    header = customtkinter.CTkLabel(main_frame, text="Borrow", font=("Poppins", 29, 'bold'),
                                    bg_color='#673F80', fg_color='#673F80', width=199, height=55)

    header.place(x=488, y=21)
    frameline = customtkinter.CTkFrame(main_frame, border_width=1, corner_radius=1, width=1060, height=1,
                                       fg_color="white", bg_color="#673F80")
    frameline.place(x=33, y=78)
    # ===================main board====================
    board = customtkinter.CTkFrame(main_frame, fg_color='#412948', bg_color="#673F80", border_width=0,
                                        corner_radius=(15), width=974, height=561)
    board.place(x=79, y=116)

    def borrow():
        rnum = 0
        rnum2 = 0
        row = sh1.max_row
        row2 = sh2.max_row
        for i in range(row):
            if input11.get() == sh1.cell(i + 1, 3).value:
                rnum = i + 1
                break
        for i in range(row2):
            print(str(sh2.cell(i + 1, 10).value))
            if "yes" !=str(sh2.cell(i + 1, 10).value) and str(sh2.cell(i + 1, 4).value)==input11.get():
                rnum2 = i + 1
                break
        if (rnum != 0) and (rnum2 == 0) and (int(input15.get()) <= 5 and int(input15.get()) >= 1):
            num = str(sh2.max_row) + "."
            fnameof = sh1.cell(rnum, 1).value
            lnameof = sh1.cell(rnum, 2).value
            phoneN = sh1.cell(rnum, 4).value
            currentDT = datetime.datetime.now()
            row2 = sh2.max_row
            loandata = [num, fnameof, lnameof, input11.get(), gradeval.get(), input13.get(),
                        input14.get(),
                        str(currentDT.date()), str(currentDT.year) + "-" + str(currentDT.month) + "-" + str(
                    currentDT.day + int(input15.get()))]
            row2 = sh2.max_row
            for i in range(len(loandata)):
                sh2.cell(row2 + 1, i + 1, value=loandata[i])
            loaddata = [input11.get(),input15.get(),str(currentDT.day)]
            for i in range(len(loaddata)):
                sheet1.cell(numberofrow+1, i + 1, value=loaddata[i])
            wb.save("tabledata.xlsx")
            currentDT = datetime.datetime.now()
            lable1 = customtkinter.CTkLabel(main_frame, text="The proccesse is successfully completed",
                                                 font=("Poppins", 18, "normal"), width=474,
                                                 height=38, bg_color="#412948", fg_color="#412948",
                                                 text_color="#F39C12")
            input11.delete(0, 10)
            input11.insert(0, "ID")
            input15.delete(0, 30)
            input15.insert(0, "Length Of Time (in day)")
            lable1.place(x=467, y=116 + 493)
            clien = Client(account_sid, auth_token)
            massage = clien.messages.create(
                body="Hi " + fnameof + " " + lnameof + ", you have succesfully borrowed grade " + gradeval.get() + " " + input14.get() + " " + input13.get() + " on " + str(
                    currentDT.date()) + " you have to return the book on " + str(currentDT.year) + "-" + str(
                    currentDT.month) + "-" + str(currentDT.day + int(input15.get())),
                from_="+13343731074",
                to="+251947853943"
            )
            input11.delete(0, 100)
            input11.insert(0, "ID")
        else:
            if rnum == 0:
                input11.configure(text_color="red")
                input11.delete(0, 10)
                input11.insert(0, "This user is not belong to the system")
                print("this user not belong the system")
            elif rnum2 != 0:
                lable2 = customtkinter.CTkLabel(main_frame, text="The user has taken a book before",
                                                font=("Poppins", 18, "normal"), width=474,
                                                height=38, bg_color="#412948", fg_color="#412948",
                                                text_color="#F39C12")
                lable2.place(x=467, y=116 + 493)
            else:
                input15.configure(text_color="red")
                input15.delete(0, 10)
                input15.insert(0, "Beyond the maximum time or less")
                print("beyond the maximum time")

    def unsave():
        input11.delete(0, 10)
        input11.insert(0, "ID")
        input15.delete(0, 30)
        input15.insert(0, "Length Of Time (in day)")


    def lables(texts, widhts, px, py):
        lable1 = customtkinter.CTkLabel(main_frame, text=texts, font=("Poppins", 22, "normal"), width=widhts,
                                             height=38, bg_color="#412948", fg_color="#412948")
        lable1.place(x=px, y=py)

    lables("Enter your ID:", 150, 301, 169)
    lables("Select Grade Of The Book:", 296, 155, 245)
    lables("Select Subject Of The Book:", 311, 140, 314)
    lables("Select Type Of The Book:", 279, 172, 390)
    lables("Enter Length Of Time:", 242, 209, 473)

    def click1(event):
        input11.delete(0, 10)

    def click5(event):
        input15.delete(0, 30)

    varinput = customtkinter.StringVar()
    input11 = customtkinter.CTkEntry(main_frame, width=459, height=44, fg_color="#412948",
                                          bg_color="#412948",
                                          border_width=1, border_color="white", placeholder_text="ID",
                                          textvariable=varinput)
    input11.insert(0, "ID")
    input11.bind("<Button-1>", click1)
    input11.place(x=474, y=166)

    grades = ["9", "10", "11", "12"]
    gradeval = customtkinter.CTkComboBox(main_frame, values=grades, width=459, height=44, fg_color="#412948",
                                              bg_color="#412948", border_width=1, border_color="white")
    gradeval.place(x=474, y=242)

    subjects = ["Mathimatics", "English", "Physics", "Chemistry", "Biology", "Afaan Oromo"]
    input13 = customtkinter.CTkComboBox(main_frame, values=subjects, width=459, height=44, fg_color="#412948",
                                             bg_color="#412948", border_width=1, border_color="white")
    input13.place(x=474, y=318)

    type = ["Extreme", "Solar", "Furtu", "TOP"]
    input14 = customtkinter.CTkComboBox(main_frame, values=type, width=459, height=44, fg_color="#412948",
                                             bg_color="#412948", border_width=1, border_color="white")
    input14.place(x=474, y=394)
    varinput = customtkinter.StringVar()
    input15 = customtkinter.CTkEntry(main_frame, width=459, height=44, fg_color="#412948",
                                          bg_color="#412948",
                                          border_width=1, border_color="white",
                                          placeholder_text="Length Of Time (in day)",
                                          textvariable=varinput)
    input15.insert(0, "Length Of Time (in day)")
    input15.bind("<Button-1>", click5)
    input15.place(x=474, y=470)

    save = customtkinter.CTkButton(main_frame, text='Save', bg_color='#412948', fg_color='#1D9321',
                                        font=("Poppins", 20, "normal"),
                                        cursor='hand2', hover_color="#0B7515", command=borrow)
    save.place(x=543, y=116 + 427)
    discard = customtkinter.CTkButton(main_frame, text='Discard', bg_color='#412948', fg_color='#D80A12',
                                           font=("Poppins", 20, "normal"),
                                           cursor='hand2', hover_color='#C10A11', command=unsave)
    discard.place(x=739, y=116 + 427)
